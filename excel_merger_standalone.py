#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel合并工具 - 独立版本（无pandas依赖）
使用内置库实现，避免打包问题
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import os
import csv
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook

class ExcelMergerStandalone:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel合并工具")
        self.root.geometry("600x500")
        self.root.resizable(False, False)

        # 设置窗口居中
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - 300
        y = (self.root.winfo_screenheight() // 2) - 250
        self.root.geometry(f"600x500+{x}+{y}")

        self.files = []
        self.setup_ui()

    def setup_ui(self):
        # 标题
        tk.Label(
            self.root,
            text="Excel 文件合并工具",
            font=("Arial", 18, "bold"),
            bg="#2196F3",
            fg="white",
            pady=15
        ).pack(fill="x")

        # 说明
        tk.Label(
            self.root,
            text="选择多个列名相同的Excel文件进行合并",
            font=("Arial", 10),
            pady=10
        ).pack()

        # 按钮框架
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame,
            text="选择文件",
            command=self.select_files,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10),
            padx=20,
            pady=5
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="清空列表",
            command=self.clear_files,
            bg="#f44336",
            fg="white",
            font=("Arial", 10),
            padx=20,
            pady=5
        ).pack(side="left", padx=5)

        # 文件列表
        list_frame = tk.Frame(self.root)
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")

        self.listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=("Arial", 10),
            height=12
        )
        self.listbox.pack(fill="both", expand=True)
        scrollbar.config(command=self.listbox.yview)

        # 状态标签
        self.status = tk.Label(
            self.root,
            text="请选择要合并的Excel文件",
            font=("Arial", 10),
            fg="#666"
        )
        self.status.pack(pady=5)

        # 合并按钮
        self.merge_btn = tk.Button(
            self.root,
            text="开始合并",
            command=self.merge_files,
            bg="#2196F3",
            fg="white",
            font=("Arial", 12, "bold"),
            padx=30,
            pady=10,
            state="disabled"
        )
        self.merge_btn.pack(pady=15)

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )

        for file in files:
            if file not in self.files:
                self.files.append(file)
                self.listbox.insert(tk.END, os.path.basename(file))

        self.update_status()

    def clear_files(self):
        self.files = []
        self.listbox.delete(0, tk.END)
        self.update_status()

    def update_status(self):
        count = len(self.files)
        if count == 0:
            self.status.config(text="请选择要合并的Excel文件")
            self.merge_btn.config(state="disabled")
        elif count == 1:
            self.status.config(text="至少需要选择2个文件")
            self.merge_btn.config(state="disabled")
        else:
            self.status.config(text=f"已选择 {count} 个文件")
            self.merge_btn.config(state="normal")

    def read_excel_file(self, filepath):
        """读取Excel文件，返回数据列表"""
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active

        data = []
        for row in ws.iter_rows(values_only=True):
            # 过滤掉全空的行
            if any(cell is not None for cell in row):
                data.append(list(row))

        wb.close()
        return data

    def write_excel_file(self, filepath, data):
        """将数据写入Excel文件"""
        wb = Workbook()
        ws = wb.active

        for row_data in data:
            ws.append(row_data)

        wb.save(filepath)
        wb.close()

    def merge_files(self):
        if len(self.files) < 2:
            messagebox.showwarning("警告", "请选择至少2个文件")
            return

        try:
            # 读取所有文件
            self.status.config(text="正在读取文件...")
            self.root.update()

            all_data = []
            first_headers = None

            for i, file in enumerate(self.files):
                self.status.config(text=f"读取 {i+1}/{len(self.files)}: {os.path.basename(file)}")
                self.root.update()

                data = self.read_excel_file(file)

                if not data:
                    messagebox.showwarning("警告", f"文件 {os.path.basename(file)} 是空的")
                    continue

                headers = data[0]  # 第一行作为列名

                if first_headers is None:
                    first_headers = headers
                    all_data.append(headers)  # 添加表头
                elif headers != first_headers:
                    # 列名不一致
                    messagebox.showerror(
                        "错误",
                        f"文件列名不一致:\n{os.path.basename(file)}\n\n"
                        f"期望: {first_headers[:5]}...\n"
                        f"实际: {headers[:5]}..."
                    )
                    self.status.config(text="合并失败：列名不一致")
                    return

                # 添加数据行（跳过表头）
                all_data.extend(data[1:])

            if not all_data or len(all_data) <= 1:
                messagebox.showwarning("警告", "没有有效的数据可以合并")
                self.status.config(text="合并失败：无有效数据")
                return

            # 合并数据
            self.status.config(text="正在合并数据...")
            self.root.update()

            # 保存文件
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")]
            )

            if save_path:
                self.status.config(text="正在保存文件...")
                self.root.update()

                self.write_excel_file(save_path, all_data)

                messagebox.showinfo(
                    "成功",
                    f"合并完成！\n共 {len(all_data)-1} 行数据（不含表头）\n保存至: {save_path}"
                )
                self.status.config(text=f"合并成功：{len(all_data)-1} 行数据")
            else:
                self.status.config(text="已取消保存")

        except Exception as e:
            messagebox.showerror("错误", f"处理文件时发生错误:\n{str(e)}")
            self.status.config(text="合并失败")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ExcelMergerStandalone()
    app.run()