#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel合并工具 - 专业版（支持图片）
支持合并包含图片的Excel文件
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

class ExcelMergerPro:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel合并工具 Pro - 支持图片")
        self.root.geometry("700x600")
        self.root.resizable(False, False)

        # 设置窗口居中
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - 350
        y = (self.root.winfo_screenheight() // 2) - 300
        self.root.geometry(f"700x600+{x}+{y}")

        self.files = []
        self.preserve_images = tk.BooleanVar(value=True)
        self.setup_ui()

    def setup_ui(self):
        # 标题
        tk.Label(
            self.root,
            text="Excel 文件合并工具 Pro",
            font=("Arial", 20, "bold"),
            bg="#2196F3",
            fg="white",
            pady=15
        ).pack(fill="x")

        # 说明
        tk.Label(
            self.root,
            text="支持合并包含图片的Excel文件",
            font=("Arial", 11),
            pady=10
        ).pack()

        # 选项框架
        option_frame = tk.Frame(self.root)
        option_frame.pack(pady=5)

        tk.Checkbutton(
            option_frame,
            text="保留图片（可能增加处理时间）",
            variable=self.preserve_images,
            font=("Arial", 10)
        ).pack()

        # 按钮框架
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame,
            text="选择文件",
            command=self.select_files,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10),
            padx=20,
            pady=5
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="清空列表",
            command=self.clear_files,
            bg="#f44336",
            fg="white",
            font=("Arial", 10),
            padx=20,
            pady=5
        ).pack(side="left", padx=5)

        # 文件列表
        list_frame = tk.Frame(self.root)
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # 创建Treeview来显示文件信息
        columns = ("文件名", "行数", "有图片")
        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)

        # 定义列
        self.tree.heading("文件名", text="文件名")
        self.tree.heading("行数", text="行数")
        self.tree.heading("有图片", text="包含图片")

        # 设置列宽
        self.tree.column("文件名", width=400)
        self.tree.column("行数", width=100)
        self.tree.column("有图片", width=100)

        # 添加滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 进度条
        self.progress = ttk.Progressbar(
            self.root,
            mode='determinate',
            length=660
        )
        self.progress.pack(pady=10, padx=20)

        # 状态标签
        self.status = tk.Label(
            self.root,
            text="请选择要合并的Excel文件",
            font=("Arial", 10),
            fg="#666"
        )
        self.status.pack(pady=5)

        # 合并按钮
        self.merge_btn = tk.Button(
            self.root,
            text="开始合并",
            command=self.merge_files,
            bg="#2196F3",
            fg="white",
            font=("Arial", 12, "bold"),
            padx=30,
            pady=10,
            state="disabled"
        )
        self.merge_btn.pack(pady=15)

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )

        for file in files:
            if file not in self.files:
                self.files.append(file)
                # 快速分析文件
                self.analyze_and_add_file(file)

        self.update_status()

    def analyze_and_add_file(self, filepath):
        """分析文件并添加到列表"""
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            row_count = ws.max_row

            # 检查是否有图片
            has_images = False
            try:
                # 尝试以读写模式打开检查图片
                wb_check = load_workbook(filepath)
                ws_check = wb_check.active
                has_images = len(ws_check._images) > 0 if hasattr(ws_check, '_images') else False
                wb_check.close()
            except:
                pass

            wb.close()

            # 添加到树形视图
            self.tree.insert("", "end", values=(
                os.path.basename(filepath),
                f"{row_count} 行",
                "是" if has_images else "否"
            ))
        except Exception as e:
            messagebox.showwarning("警告", f"无法分析文件 {os.path.basename(filepath)}: {str(e)}")

    def clear_files(self):
        self.files = []
        # 清空树形视图
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.update_status()
        self.progress['value'] = 0

    def update_status(self):
        count = len(self.files)
        if count == 0:
            self.status.config(text="请选择要合并的Excel文件")
            self.merge_btn.config(state="disabled")
        elif count == 1:
            self.status.config(text="至少需要选择2个文件")
            self.merge_btn.config(state="disabled")
        else:
            self.status.config(text=f"已选择 {count} 个文件，准备合并")
            self.merge_btn.config(state="normal")

    def copy_images(self, source_ws, target_ws, row_offset):
        """复制图片从源工作表到目标工作表"""
        if not self.preserve_images.get():
            return

        try:
            if hasattr(source_ws, '_images'):
                for img in source_ws._images:
                    # 获取图片位置
                    row = img.anchor._from.row
                    col = img.anchor._from.col

                    # 创建新图片对象
                    new_img = Image(BytesIO(img._data()))

                    # 设置新位置（加上行偏移）
                    new_img.anchor = f"{openpyxl.utils.get_column_letter(col + 1)}{row + row_offset + 1}"

                    # 保持原始大小
                    new_img.width = img.width
                    new_img.height = img.height

                    # 添加到目标工作表
                    target_ws.add_image(new_img)
        except Exception as e:
            print(f"复制图片时出错: {e}")

    def merge_files(self):
        if len(self.files) < 2:
            messagebox.showwarning("警告", "请选择至少2个文件")
            return

        try:
            # 初始化进度条
            self.progress['value'] = 0
            self.progress['maximum'] = len(self.files) + 1

            # 创建新工作簿
            merged_wb = Workbook()
            merged_ws = merged_wb.active

            first_headers = None
            current_row = 0
            total_data_rows = 0
            image_count = 0

            for i, file in enumerate(self.files):
                self.status.config(text=f"处理文件 {i+1}/{len(self.files)}: {os.path.basename(file)}")
                self.root.update()

                # 打开源文件
                source_wb = load_workbook(file, data_only=False)
                source_ws = source_wb.active

                # 获取数据
                data = []
                for row in source_ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        data.append(list(row))

                if not data:
                    source_wb.close()
                    continue

                headers = data[0]

                if first_headers is None:
                    # 第一个文件，复制表头和列宽
                    first_headers = headers
                    merged_ws.append(headers)
                    current_row = 1

                    # 复制列宽
                    for col_idx, col in enumerate(source_ws.iter_cols(), 1):
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        if source_ws.column_dimensions[col_letter].width:
                            merged_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

                    # 设置行高（特别是对于包含图片的行）
                    for row_idx in range(1, len(data) + 1):
                        if source_ws.row_dimensions[row_idx].height:
                            merged_ws.row_dimensions[current_row + row_idx - 1].height = source_ws.row_dimensions[row_idx].height

                elif headers != first_headers:
                    source_wb.close()
                    messagebox.showerror(
                        "错误",
                        f"文件列名不一致:\n{os.path.basename(file)}"
                    )
                    self.status.config(text="合并失败：列名不一致")
                    return

                # 添加数据行
                for row_data in data[1:]:
                    merged_ws.append(row_data)
                    current_row += 1
                    total_data_rows += 1

                # 复制图片（如果选择保留）
                if self.preserve_images.get():
                    self.status.config(text=f"复制图片: {os.path.basename(file)}")
                    self.root.update()

                    # 计算行偏移（第一个文件偏移0，后续文件需要偏移）
                    row_offset = current_row - len(data)
                    if i == 0:
                        row_offset = 0

                    # 复制图片
                    self.copy_images(source_ws, merged_ws, row_offset)

                    # 统计图片数量
                    if hasattr(source_ws, '_images'):
                        image_count += len(source_ws._images)

                source_wb.close()

                # 更新进度条
                self.progress['value'] = i + 1
                self.root.update()

            # 保存文件
            self.status.config(text="准备保存...")
            self.root.update()

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialfile="merged_result.xlsx"
            )

            if save_path:
                self.status.config(text="正在保存文件...")
                self.root.update()

                merged_wb.save(save_path)
                merged_wb.close()

                # 完成进度条
                self.progress['value'] = self.progress['maximum']

                image_info = f"\n包含 {image_count} 张图片" if image_count > 0 else ""
                messagebox.showinfo(
                    "成功",
                    f"合并完成！\n\n"
                    f"合并了 {len(self.files)} 个文件\n"
                    f"共 {total_data_rows} 行数据{image_info}\n"
                    f"保存至: {save_path}"
                )
                self.status.config(text=f"合并成功：{total_data_rows} 行数据")
            else:
                merged_wb.close()
                self.status.config(text="已取消保存")

        except Exception as e:
            messagebox.showerror("错误", f"处理文件时发生错误:\n{str(e)}")
            self.status.config(text="合并失败")
        finally:
            self.progress['value'] = 0

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ExcelMergerPro()
    app.run()